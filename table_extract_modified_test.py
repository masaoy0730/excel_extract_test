from openpyxl import load_workbook

# Excelファイルを読み込む
file_path = "<filename>"  # 読み込むExcelファイルのパス
workbook = load_workbook(file_path)
sheet = workbook.active  # アクティブなシートを選択

# 列名（1行目）を取得
column_names = [cell.value for cell in sheet[1]]  # 1行目のセル値を取得

# 2行目以降のデータを列名とセットにして配列を作成
text_data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # 2行目以降のデータを取得
    row_data = {column_names[i]: value for i, value in enumerate(row)}  # 列名と値をセットにする
    text_data.append(str(row_data))  # 辞書形式を文字列に変換して配列に追加

# 結果を出力
print(text_data[0]) #配列なので、出力を確認するときはインデックスを指定する
