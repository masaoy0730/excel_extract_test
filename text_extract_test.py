# Excelファイルのテキスト抽出
pip install openpyxl

# OpenPyXLはExcelファイル（.xlsx形式）を読み書きするためのライブラリ
import openpyxl

# Excelファイルのテキスト抽出の関数
def extract_text_from_excel(filepath):
    # Excelファイルを開く
    workbook = openpyxl.load_workbook(filepath)
    text_data = []

    # 各シートを処理
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    text_data.append(str(cell.value))

    return "\n".join(text_data)

# 実行例
filepath = "<file name>.xlsx"
print(extract_text_from_excel(filepath))
